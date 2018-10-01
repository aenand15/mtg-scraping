from selenium import webdriver
from selenium.webdriver.firefox.options import Options
import openpyxl
import time
import requests
import sys

def createDriver():
	options = Options()
	options.set_headless(headless=True)
	driver = webdriver.Firefox(firefox_options=options)
	return driver

def getcardPrice(cardtitle, edition, driver):
	url = 'https://www.cardkingdom.com/purchasing/mtg_singles/?filter%5bsort%5D=price_desc&filter%5bsearch%5D=mtg_advanced&filter%5bname%5D=' + cardtitle +'&filter%5bcategory_id%5D=' + edition + '&filter%5bnonfoil%5D=1&filter%5bprice_op%5D=&filter%5bprice%5D=&filter%5bnonfoil%5D=1&filter%5bprice_op%5D=&filter%5bprice%5D='
	driver.get(url)
	try: 
		price_box = driver.find_element_by_class_name(name="sellDollarAmount")
		price = price_box.text
	except:
		price = "missing"
	print(price)
	return price

def createHash():
	editions = {
		"standard": 2779,
		"modern": 2864,
		"3rd e": 2345,
		"4th e": 2350,
		"5th e": 2355,
		"6th e": 2360,
		"7th e": 2365,
		"8th e": 2370,
		"9th e": 2375,
		"10th e": 2380,
		"2010 core set": 2789,
		"2011 core set": 2847,
		"2012 core set": 2863,
		"2013 core set": 2876,
		"2014 core set": 2895,
		"2015 core set": 2910,
		"aether revolt": 3030,
		"alara reborn": 2385,
		"alliances": 2390,
		"alpha": 2395,
		"amonkhet": 3042,
		"anthologies": 2400,
		"antiquities": 2405,
		"apocalypse": 2410,
		"arabian nights": 2415,
		"archenemy": 2846,
		"archenemy nicol bo": 3048,
		"avacyn restored": 2874,
		"battle for zendikar": 2953,
		"battle royale": 2420,
		"battlebond": 3088,
		"beatdown": 2425,
		"beta": 2430,
		"betrayers of kamigaw": 2435,
		"born of the gods": 2903,
		"champions of kamigawa": 2440,
		"chronicles": 2445,
		"coldsnap": 2450,
		"coldsnap theme decks": 3099,
		"collectors ed": 2455,
		"collectors ed intl": 2460,
		"commander": 2862,
		"commander 2013": 2902,
		"commander 2014": 2916,
		"commander 2015": 2958,
		"commander 2016": 2949,
		"commander 2017": 3055,
		"commander 2018": 3100,
		"commander anthology": 3047,
		"commander anthology vol 2": 3089,
		"commander's arsenal": 2888,
		"conflux": 2783,
		"conspiracy": 2908,
		"conspiracy take the crown": 2977,
		"core set 2019": 3097,
		"dark ascension": 2870,
		"darksteel": 2465,
		"deckmaster": 2470,
		"dissension": 2475,
		"dominaria": 3086,
		"dragon's maze": 2892,
		"dragons of tarkir": 2938,
		"duel decks ajani vs. nicol bolas": 2865,
		"dd anthology": 2918,
		"dd blessed/cursed": 2969,
		"dd divine/demonic": 2480,
		"dd elspeth/kiora": 2936,
		"dd elspeth/tezzeret": 2851,
		"dd elves/goblins": 2485,
		"dd elves/inventor": 3084,
		"dd garruk/liliana": 2838,
		"dd heroes/monsters": 2896,
		"dd izzet/golgari": 2878,
		"dd jace/chandra": 2490,
		"dd jace/vraska": 2904,
		"dd knights/dragons": 2860,
		"dd merfolk/goblins": 3062,
		"dd mind/might": 3041,
		"dd nissa/ob nixilis": 2980,
		"dd phyrexia/the coal": 2841,
		"dd sorin/tibalt": 2891,
		"dd speed/cunning": 2911,
		"dd venser/koth": 2873,
		"dd zendikar/eldrazi": 2951,
		"duels of the planeswalker": 2845,
		"eldritch moon": 2976,
		"eternal masters": 2973,
		"eventide": 2495,
		"exodus": 2500,
		"explorers of ixalan": 3064,
		"fallen empire": 2505,
		"fate reforged": 2923,
		"fifth dawn": 2510,
		"ftv angels": 2952,
		"ftv annihilation": 2913,
		"ftv dragons": 2515,
		"ftv exiled": 2815,
		"ftv legends": 2868,
		"ftv lore": 2979,
		"ftv realms": 2883,
		"ftv relics": 2850,
		"ftv transform": 3065,
		"ftv twenty": 2899,
		"future sight": 2520,
		"gatecrash": 2890,
		"guildpact": 2525,
		"homelands": 2530,
		"hour of devastation": 3051,
		"ice age": 2535,
		"iconic masters": 3059,
		"innistrad": 2866,
		"invasion": 2540,
		"ixalan": 3058,
		"journey into nyx": 2905,
		"judgement": 2545,
		"kaladesh": 2983,
		"khans of tarkir": 2914,
		"legends": 2550,
		"legions": 2555,
		"lorwyn": 2560,
		"magic origins": 2950,
		"mps expeditions": 2960,
		"mps inventions": 2984,
		"mps: invocations": 3044,
		"masters 25": 3078,
		"mercadian masques": 2565,
		"mirage": 2570,
		"mirrodin": 2575,
		"mirrodin besieged": 2859,
		"modern event deck": 2907,
		"modern masters": 2894,
		"modern masters 2015": 2947,
		"modern masters 2017": 3032,
		"morningtide": 2580,
		"nemesis": 2590,
		"new phyrexia": 2861,
		"oath of the gatewatch": 2967,
		"odyssey": 2595,
		"onslaught": 2600,
		"planar chaos": 2605,
		"planechase": 2839,
		"planechase 2012": 2875,
		"planechase anthology": 2989,
		"planeshift": 2610,
		"portal": 2615,
		"portal 3k": 2620,
		"portal 2": 2625,
		"pds fire and light": 2854,
		"pds graveborn": 2867,
		"pds slivers": 2837,
		"promotional": 2630,
		"prophecy": 2635,
		"ravnica": 2640,
		"return to ravnica": 2884,
		"rise of the eldrazi": 2843,
		"rivals of ixalan": 3076,
		"saviors of kamigawa": 2645,
		"scars of mirrodin": 2852,
		"scourge": 2650,
		"shadowmoor": 2655,
		"shadows over innistrad": 2971,
		"shards of alara": 2660,
		"signature spellbook": 3091,
		"starter 1999": 2670,
		"strongholds": 2680,
		"tempest": 2685,
		"dark": 2690,
		"theros": 2900,
		"time spiral": 2695,
		"timeshifted": 2700,
		"torment": 2705,
		"unglued": 2710,
		"unhinged": 2715,
		"unlimited": 2720,
		"unstable": 3075,
		"urza's destiny": 2725,
		"urza's legacy": 2730,
		"urza's saga": 2735,
		"vanguard": 2740,
		"visions": 2745,
		"weatherlights": 2750,
		"world championships": 2975,
		"worldwake": 2840,
		"zendikar": 2826,
	}
	return editions

def main():
	start_time = time.time()
	driver = createDriver()
	editions = createHash()
	filename = "EXCEL FILENAME"
	wb = openpyxl.load_workbook(filename)
	sheet = wb['SHEET NAME']
	completed = 0
	i = STARTING LINE
	fCounter = 0
	while completed < NUM TO COMPLETE:
		cardName = sheet.cell(row=i, column=1).value
		cardEdition = sheet.cell(row=i, column=3).value
		cardEdition = cardEdition.lower()
		print(cardName)
		print(cardEdition)
		price = getcardPrice(cardName, str(editions[cardEdition]), driver)
		if 'failure' in price:
			fCounter = fCounter + 1
		sheet.cell(row=i, column=4).value = price
		if completed % 10 == 0:
			wb.save(filename)
		print(completed)
		completed = completed + 1
		i = i + 1
		time.sleep(3)
	report = "Number completed: %d Number failed: %d --- Completed in %s Seconds ---" % ((completed - fCounter), fCounter, (time.time() - start_time))
	sheet.cell(row=i+1, column=5).value = report
	wb.save(filename)
	print(report)
	driver.quit()

if __name__ == "__main__":
	main() 
